"""
generate_excel.py - Gera cheatsheet_excel.xlsx
================================================
Cria um arquivo Excel completo com multiplas abas, dados reais
e formulas Excel prontas para estudo.

Abas geradas:
  Dados          - dataset employees raw + formatacao
  Filtros        - exemplos de filtros com formulas
  Agrupamento    - SOMASE, CONT.SE, MEDIASE por departamento
  Datas          - ANO, MES, DIA, HOJE, calculos de tempo
  Condicional    - SE, IFS, SEERRO, coluna de nivel
  Lookup         - PROCV, PROCX, INDICE+CORRESP
  Metricas       - MIN, MAX, ORDEM.EQ, GRANDE, SOMARPRODUTO

Execute: python docs/examples/excel/generate_excel.py
Dataset: datasets/employees.csv, datasets/customers.csv, datasets/orders.csv
Saida:   docs/examples/excel/cheatsheet_excel.xlsx
"""

import os
import csv
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
except ImportError:
    print("ERRO: openpyxl nao encontrado. Execute: pip install openpyxl")
    raise

# ── Caminhos ─────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT   = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", ".."))
DATASETS    = os.path.join(REPO_ROOT, "datasets")
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "cheatsheet_excel.xlsx")


# ── Estilos ──────────────────────────────────────────────────────────────────
def header_style(ws, row, col, texto, cor_hex="1F4E79"):
    cell = ws.cell(row=row, column=col, value=texto)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor=cor_hex)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def section_style(ws, row, col, texto, cor_hex="2E75B6"):
    cell = ws.cell(row=row, column=col, value=texto)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor=cor_hex)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return cell


def formula_style(ws, row, col, formula, label=""):
    """Coloca formula e destaca a celula."""
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = PatternFill("solid", fgColor="FFF2CC")  # amarelo claro
    cell.font = Font(color="000000", size=10)
    cell.alignment = Alignment(horizontal="left")
    if label:
        ws.cell(row=row, column=col - 1, value=label).font = Font(italic=True, size=10)
    return cell


def thin_border():
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def apply_table_border(ws, min_row, max_row, min_col, max_col):
    thin = thin_border()
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                            min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin


# ── Leitura dos datasets ─────────────────────────────────────────────────────
def read_csv(nome):
    path = os.path.join(DATASETS, nome)
    rows = []
    with open(path, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)
    return rows


# ── ABA 1: Dados ─────────────────────────────────────────────────────────────
def create_dados(wb):
    ws = wb.active
    ws.title = "Dados"
    ws.sheet_view.showGridLines = True

    # Titulo
    ws.merge_cells("A1:H1")
    cell = ws["A1"]
    cell.value = "Dataset: Funcionarios (employees.csv)"
    cell.font = Font(bold=True, size=14, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabecalhos
    headers = ["ID", "Nome", "Departamento", "Cargo", "Cidade", "Salario", "Data Admissao", "Ativo"]
    cols_w  = [6,    22,     16,              20,      16,        12,       16,              8]

    for i, (h, w) in enumerate(zip(headers, cols_w), start=1):
        header_style(ws, 2, i, h)
        ws.column_dimensions[get_column_letter(i)].width = w

    # Dados
    emp = read_csv("employees.csv")
    cores = ["F2F7FF", "FFFFFF"]  # alternado
    for r, row in enumerate(emp, start=3):
        cor = cores[r % 2]
        fill = PatternFill("solid", fgColor=cor)
        vals = [
            row.get("id",""),
            row.get("nome",""),
            row.get("departamento",""),
            row.get("cargo",""),
            row.get("cidade",""),
            float(row["salario"]) if row.get("salario") else None,
            row.get("data_admissao",""),
            row.get("ativo",""),
        ]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")
            if c == 6 and v:  # salario
                cell.number_format = 'R$ #,##0.00'
            if c == 7 and v:  # data
                cell.number_format = 'DD/MM/YYYY'

    n_linhas = len(emp) + 2
    apply_table_border(ws, 2, n_linhas, 1, 8)

    # Painel de resumo rapido
    ws.merge_cells("J2:M2")
    section_style(ws, 2, 10, "Resumo Rapido")
    ws.column_dimensions["J"].width = 24
    ws.column_dimensions["K"].width = 16
    ws.column_dimensions["L"].width = 4
    ws.column_dimensions["M"].width = 16

    resumo = [
        ("Total de funcionarios",   "=CONT.VALORES(A3:A54)"),
        ("Total de funcionarios ativos", '=CONT.SE(H3:H54,"Sim")'),
        ("Salario medio",           "=MEDIA(F3:F54)"),
        ("Maior salario",           "=MAXIMO(F3:F54)"),
        ("Menor salario",           "=MINIMO(F3:F54)"),
        ("Folha total",             "=SOMA(F3:F54)"),
        ("Nulos em salario",        "=CONT.VALORES(F3:F54)-CONT.NUM(F3:F54)"),
    ]
    for i, (label, formula) in enumerate(resumo, start=3):
        ws.cell(row=i, column=10, value=label).font = Font(size=10)
        cell = ws.cell(row=i, column=11, value=formula)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(bold=True)
        if "MEDIA" in formula or "MAXIMO" in formula or "MINIMO" in formula or "SOMA" in formula:
            cell.number_format = 'R$ #,##0.00'

    ws.row_dimensions[2].height = 20
    print("  [OK] Aba 'Dados' criada")
    return n_linhas


# ── ABA 2: Filtros ───────────────────────────────────────────────────────────
def create_filtros(wb):
    ws = wb.create_sheet("Filtros")

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Exemplos de Filtros com Formulas Excel"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 40

    # Coluna de referencia
    section_style(ws, 2, 1, "Descricao")
    section_style(ws, 2, 2, "Formula Excel")
    section_style(ws, 2, 5, "Descricao")
    section_style(ws, 2, 6, "Formula Excel")

    filtros_esq = [
        ("Contar TI",               '=CONT.SE(Dados!C3:C54,"TI")'),
        ("Contar ativos",           '=CONT.SE(Dados!H3:H54,"Sim")'),
        ("Contar salario > 8000",   "=CONT.SE(Dados!F3:F54,\">8000\")"),
        ("Contar nulos salario",    "=CONT.VALORES(Dados!F3:F54)-CONT.NUM(Dados!F3:F54)"),
        ("Soma salario TI",         '=SOMASE(Dados!C3:C54,"TI",Dados!F3:F54)'),
        ("Soma salario > 5000",     '=SOMASE(Dados!F3:F54,">5000",Dados!F3:F54)'),
        ("Media TI",                '=MEDIASE(Dados!C3:C54,"TI",Dados!F3:F54)'),
        ("Valor min salario",       "=MINIMO(Dados!F3:F54)"),
        ("Valor max salario",       "=MAXIMO(Dados!F3:F54)"),
        ("Total geral folha",       "=SOMA(Dados!F3:F54)"),
    ]
    filtros_dir = [
        ("CONT.SES - TI E Ativo",    '=CONT.SES(Dados!C3:C54,"TI",Dados!H3:H54,"Sim")'),
        ("SOMASES - Vendas E Ativo",  '=SOMASES(Dados!F3:F54,Dados!C3:C54,"Vendas",Dados!H3:H54,"Sim")'),
        ("MEDIASES - TI E > 5000",    '=MEDIASES(Dados!F3:F54,Dados!C3:C54,"TI",Dados!F3:F54,">5000")'),
        ("Nome maior salario",        "=INDICE(Dados!B3:B54,CORRESP(MAXIMO(Dados!F3:F54),Dados!F3:F54,0))"),
        ("Nome menor salario",        "=INDICE(Dados!B3:B54,CORRESP(MINIMO(Dados!F3:F54),Dados!F3:F54,0))"),
        ("3o maior salario",          "=GRANDE(Dados!F3:F54,3)"),
        ("3o menor salario",          "=PEQUENO(Dados!F3:F54,3)"),
        ("Unicos departamentos",      "=SOMARPRODUTO(1/CONT.SE(Dados!C3:C53,Dados!C3:C53))"),
        ("% salario > 8000",         "=CONT.SE(Dados!F3:F54,\">8000\")/CONT.VALORES(Dados!F3:F54)"),
        ("Folha ativos",              '=SOMASE(Dados!H3:H54,"Sim",Dados!F3:F54)'),
    ]

    for i, (desc, form) in enumerate(filtros_esq, start=3):
        ws.cell(row=i, column=1, value=desc).font = Font(size=10)
        cell = ws.cell(row=i, column=2, value=form)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(size=10)
        if "MEDIA" in form or "SOMA" in form or "MIN" in form or "MAX" in form or "GRANDE" in form:
            cell.number_format = 'R$ #,##0.00'
        if "%" in desc:
            cell.number_format = '0.0%'

    for i, (desc, form) in enumerate(filtros_dir, start=3):
        ws.cell(row=i, column=5, value=desc).font = Font(size=10)
        cell = ws.cell(row=i, column=6, value=form)
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.font = Font(size=10)
        if "MEDIA" in form or "SOMA" in form or "GRANDE" in form or "PEQUENO" in form:
            cell.number_format = 'R$ #,##0.00'
        if "%" in desc:
            cell.number_format = '0.0%'

    apply_table_border(ws, 2, 12, 1, 2)
    apply_table_border(ws, 2, 12, 5, 6)

    # Nota
    ws.cell(row=14, column=1, value="Nota: As formulas acima referenciam a aba 'Dados'. Certifique-se de que ela existe.").font = Font(italic=True, color="595959", size=9)
    print("  [OK] Aba 'Filtros' criada")


# ── ABA 3: Agrupamento ───────────────────────────────────────────────────────
def create_agrupamento(wb):
    ws = wb.create_sheet("Agrupamento")

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Agrupamento por Departamento (equivalente ao GROUP BY)"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    departamentos = ["Financeiro", "Marketing", "Operacoes", "RH", "TI", "Vendas"]

    # Cabecalhos
    headers = ["Departamento", "Qtd Funcionarios", "Salario Total", "Salario Medio", "Maior Salario", "Menor Salario", "% da Folha"]
    widths  = [18, 18, 16, 15, 15, 15, 12]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        header_style(ws, 2, i, h, cor_hex="375623")
        ws.column_dimensions[get_column_letter(i)].width = w

    total_row = len(departamentos) + 3

    for r, dept in enumerate(departamentos, start=3):
        row_fill = PatternFill("solid", fgColor="F2F7F0" if r % 2 == 0 else "FFFFFF")

        ws.cell(row=r, column=1, value=dept).font = Font(bold=True, size=10)

        formulas = [
            f'=CONT.SE(Dados!C$3:C$54,A{r})',
            f'=SOMASE(Dados!C$3:C$54,A{r},Dados!F$3:F$54)',
            f'=MEDIASE(Dados!C$3:C$54,A{r},Dados!F$3:F$54)',
            f'=MAXSES(Dados!F$3:F$54,Dados!C$3:C$54,A{r})',
            f'=MINSES(Dados!F$3:F$54,Dados!C$3:C$54,A{r})',
            f'=C{r}/C${total_row}',
        ]
        formats = ["General", 'R$ #,##0.00', 'R$ #,##0.00', 'R$ #,##0.00', 'R$ #,##0.00', '0.00%']

        for c, (form, fmt) in enumerate(zip(formulas, formats), start=2):
            cell = ws.cell(row=r, column=c, value=form)
            cell.fill = PatternFill("solid", fgColor="FFF2CC")
            cell.number_format = fmt
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="right")

        ws.cell(row=r, column=1).fill = row_fill

    # Linha de total
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=10)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor="375623")
    ws.cell(row=total_row, column=1).font = Font(bold=True, color="FFFFFF")

    totais = [
        (2, f'=SOMA(B3:B{total_row-1})', "General"),
        (3, f'=SOMA(C3:C{total_row-1})', 'R$ #,##0.00'),
        (4, f'=MEDIA(Dados!F3:F54)',      'R$ #,##0.00'),
        (5, f'=MAXIMO(Dados!F3:F54)',     'R$ #,##0.00'),
        (6, f'=MINIMO(Dados!F3:F54)',     'R$ #,##0.00'),
        (7, "=SOMA(G3:G8)",               '0.00%'),
    ]
    for col, form, fmt in totais:
        cell = ws.cell(row=total_row, column=col, value=form)
        cell.fill = PatternFill("solid", fgColor="D9E8D4")
        cell.font = Font(bold=True, size=10)
        cell.number_format = fmt

    apply_table_border(ws, 2, total_row, 1, 7)

    # Nota sobre MAXSES/MINSES
    ws.cell(row=total_row+2, column=1,
            value="Nota: MAXSES e MINSES requerem Excel 2019 ou 365. Em versoes mais antigas, use MAXIMO com SEERRO(INDICE...)").font = Font(italic=True, color="595959", size=9)
    ws.merge_cells(f"A{total_row+2}:G{total_row+2}")

    print("  [OK] Aba 'Agrupamento' criada")


# ── ABA 4: Datas ─────────────────────────────────────────────────────────────
def create_datas(wb):
    ws = wb.create_sheet("Datas")

    ws.merge_cells("A1:F1")
    cell = ws["A1"]
    cell.value = "Trabalhar com Datas no Excel"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 3
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 38

    section_style(ws, 2, 1, "Funcao / Descricao")
    section_style(ws, 2, 2, "Formula (referencia: Dados!G3 = data_admissao)")
    section_style(ws, 2, 5, "Calculo Avancado")
    section_style(ws, 2, 6, "Formula")

    formulas_basicas = [
        ("Extrair Ano",             "=ANO(Dados!G3)"),
        ("Extrair Mes",             "=MES(Dados!G3)"),
        ("Extrair Dia",             "=DIA(Dados!G3)"),
        ("Nome do Mes",             '=TEXTO(Dados!G3,"MMMM")'),
        ("Dia da Semana (nome)",    '=TEXTO(Dados!G3,"DDDD")'),
        ("Numero dia semana",       "=DIA.DA.SEMANA(Dados!G3,2)"),
        ("Trimestre",               "=INT((MES(Dados!G3)-1)/3)+1"),
        ("Semana do ano",           "=NUM.SEMANA(Dados!G3,2)"),
        ("Hoje",                    "=HOJE()"),
        ("Agora",                   "=AGORA()"),
    ]
    formulas_avancadas = [
        ("Tempo de casa (dias)",    "=HOJE()-Dados!G3"),
        ("Tempo de casa (anos)",    "=FRAÇÃO.ANO(Dados!G3,HOJE())"),
        ("Tempo de casa (meses)",   "=DATAM(Dados!G3,0)"),
        ("Data + 30 dias",          "=Dados!G3+30"),
        ("Primeiro dia do mes",     "=DATA(ANO(Dados!G3),MES(Dados!G3),1)"),
        ("Ultimo dia do mes",       "=FIMMÊS(Dados!G3,0)"),
        ("Dias entre duas datas",   "=DIAS(HOJE(),Dados!G3)"),
        ("Dias uteis entre datas",  "=DIATRABALHO(Dados!G3,30)"),
        ("Filtrar ano 2023",        '=CONT.SE(Dados!G3:G54,">=01/01/2023")-CONT.SE(Dados!G3:G54,">=01/01/2024")'),
        ("Mais recente admissao",   "=MAXIMO(Dados!G3:G54)"),
    ]

    for i, (desc, form) in enumerate(formulas_basicas, start=3):
        ws.cell(row=i, column=1, value=desc).font = Font(size=10)
        cell = ws.cell(row=i, column=2, value=form)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(size=10)
        if "HOJE" in form and "CONT" not in form:
            cell.number_format = "DD/MM/YYYY"
        if "AGORA" in form:
            cell.number_format = "DD/MM/YYYY HH:MM"

    for i, (desc, form) in enumerate(formulas_avancadas, start=3):
        ws.cell(row=i, column=5, value=desc).font = Font(size=10)
        cell = ws.cell(row=i, column=6, value=form)
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.font = Font(size=10)
        if "data" in desc.lower() or "recente" in desc.lower():
            cell.number_format = "DD/MM/YYYY"
        if "anos" in desc.lower():
            cell.number_format = "0.0"

    apply_table_border(ws, 2, 12, 1, 2)
    apply_table_border(ws, 2, 12, 5, 6)
    print("  [OK] Aba 'Datas' criada")


# ── ABA 5: Condicional ───────────────────────────────────────────────────────
def create_condicional(wb):
    ws = wb.create_sheet("Condicional")

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Coluna Condicional — SE, IFS, SEERRO, SWITCH"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    headers = ["Nome", "Departamento", "Salario", "Nivel (SE aninhado)", "Nivel (IFS)", "Faixa Salarial", "Ativo Label"]
    widths  = [22, 16, 14, 22, 22, 18, 14]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        header_style(ws, 2, i, h, cor_hex="7030A0")
        ws.column_dimensions[get_column_letter(i)].width = w

    emp = read_csv("employees.csv")
    for r, row in enumerate(emp, start=3):
        cor = "F9F0FF" if r % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=cor)
        nome  = row.get("nome", "")
        dept  = row.get("departamento", "")
        sal   = float(row["salario"]) if row.get("salario") else None
        ativo = row.get("ativo", "")

        ws.cell(row=r, column=1, value=nome).fill = fill
        ws.cell(row=r, column=2, value=dept).fill = fill
        cell = ws.cell(row=r, column=3, value=sal)
        cell.fill = fill
        cell.number_format = 'R$ #,##0.00'

        # SE aninhado
        cell = ws.cell(row=r, column=4,
                       value=f'=SE(C{r}>12000,"Especialista",SE(C{r}>8000,"Senior",SE(C{r}>5000,"Pleno","Junior")))')
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(size=9)

        # IFS (Excel 2016+)
        cell = ws.cell(row=r, column=5,
                       value=f'=IFS(C{r}>12000,"Especialista",C{r}>8000,"Senior",C{r}>5000,"Pleno",VERDADEIRO,"Junior")')
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(size=9)

        # Faixa salarial
        cell = ws.cell(row=r, column=6,
                       value=f'=SE(C{r}>10000,"Alto",SE(C{r}>6000,"Medio","Baixo"))')
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.font = Font(size=9)

        # Ativo Label
        cell = ws.cell(row=r, column=7,
                       value=f'=SE(G{r}="Sim","Ativo","Inativo")')
        cell.fill = PatternFill("solid", fgColor="E2EFDA")
        cell.font = Font(size=9)

    n = len(emp) + 2
    apply_table_border(ws, 2, n, 1, 7)
    print("  [OK] Aba 'Condicional' criada")


# ── ABA 6: Lookup ────────────────────────────────────────────────────────────
def create_lookup(wb):
    ws = wb.create_sheet("Lookup")

    ws.merge_cells("A1:G1")
    cell = ws["A1"]
    cell.value = "Lookup — PROCV, PROCX, INDICE+CORRESP"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # Secao de exemplos
    section_style(ws, 2, 1, "ID Buscado")
    section_style(ws, 2, 2, "PROCV — Nome")
    section_style(ws, 2, 3, "PROCV — Departamento")
    section_style(ws, 2, 4, "PROCV — Salario")
    section_style(ws, 2, 5, "PROCX — Cargo")
    section_style(ws, 2, 6, "INDICE+CORRESP — Cidade")
    section_style(ws, 2, 7, "SEERRO — seguro")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 22

    # IDs para buscar (alguns existem, um nao existe para demonstrar SEERRO)
    ids = [1, 5, 10, 20, 35, 50, 99]
    for r, id_val in enumerate(ids, start=3):
        ws.cell(row=r, column=1, value=id_val).font = Font(bold=True)

        # PROCV
        ws.cell(row=r, column=2,
                value=f'=PROCV(A{r},Dados!$A$3:$H$54,2,FALSO)').fill = PatternFill("solid", fgColor="FFF2CC")
        ws.cell(row=r, column=3,
                value=f'=PROCV(A{r},Dados!$A$3:$H$54,3,FALSO)').fill = PatternFill("solid", fgColor="FFF2CC")
        cell = ws.cell(row=r, column=4,
                value=f'=PROCV(A{r},Dados!$A$3:$H$54,6,FALSO)')
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.number_format = 'R$ #,##0.00'

        # PROCX (Excel 365 / 2021)
        ws.cell(row=r, column=5,
                value=f'=PROCX(A{r},Dados!$A$3:$A$54,Dados!$D$3:$D$54,"Nao encontrado")').fill = PatternFill("solid", fgColor="E2EFDA")

        # INDICE + CORRESP
        ws.cell(row=r, column=6,
                value=f'=INDICE(Dados!$E$3:$E$54,CORRESP(A{r},Dados!$A$3:$A$54,0))').fill = PatternFill("solid", fgColor="DDEEFF")

        # SEERRO + PROCV (tratamento de erro)
        ws.cell(row=r, column=7,
                value=f'=SEERRO(PROCV(A{r},Dados!$A$3:$H$54,2,FALSO),"ID nao encontrado")').fill = PatternFill("solid", fgColor="FFE0E0")

    apply_table_border(ws, 2, 3+len(ids)-1, 1, 7)

    # Legenda
    ws.cell(row=12, column=1, value="Amarelo = PROCV  |  Verde = PROCX (Excel 365+)  |  Azul = INDICE+CORRESP  |  Vermelho = SEERRO").font = Font(italic=True, color="595959", size=9)
    ws.merge_cells("A12:G12")

    print("  [OK] Aba 'Lookup' criada")


# ── ABA 7: Metricas ──────────────────────────────────────────────────────────
def create_metricas(wb):
    ws = wb.create_sheet("Metricas")

    ws.merge_cells("A1:D1")
    cell = ws["A1"]
    cell.value = "Metricas Avancadas — Ranking, Top N, Percentual"
    cell.font = Font(bold=True, size=13, color="1F4E79")
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    headers = ["Nome", "Departamento", "Salario", "Ranking (ORDEM.EQ)"]
    widths  = [22, 16, 14, 20]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        header_style(ws, 2, i, h, cor_hex="C55A11")
        ws.column_dimensions[get_column_letter(i)].width = w

    emp = read_csv("employees.csv")
    for r, row in enumerate(emp, start=3):
        cor = "FFF4E6" if r % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=cor)
        nome = row.get("nome","")
        dept = row.get("departamento","")
        sal  = float(row["salario"]) if row.get("salario") else None

        ws.cell(row=r, column=1, value=nome).fill = fill
        ws.cell(row=r, column=2, value=dept).fill = fill
        cell = ws.cell(row=r, column=3, value=sal)
        cell.fill = fill
        cell.number_format = 'R$ #,##0.00'

        n_total = len(emp) + 2
        cell = ws.cell(row=r, column=4,
                       value=f'=ORDEM.EQ(C{r},$C$3:$C${n_total},0)')
        cell.fill = PatternFill("solid", fgColor="FFF2CC")

    n = len(emp) + 2

    # Painel de metricas
    section_style(ws, 2, 6, "Metrica", cor_hex="C55A11")
    section_style(ws, 2, 7, "Formula", cor_hex="C55A11")
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 42

    metricas = [
        ("1o maior salario",   "=GRANDE($C$3:$C$54,1)"),
        ("2o maior salario",   "=GRANDE($C$3:$C$54,2)"),
        ("3o maior salario",   "=GRANDE($C$3:$54,3)"),
        ("1o menor salario",   "=PEQUENO($C$3:$C$54,1)"),
        ("Mediana",            "=MED($C$3:$C$54)"),
        ("Desvio Padrao",      "=DESVPAD($C$3:$C$54)"),
        ("% acima da media",   "=CONT.SE($C$3:$C$54,\">\"&MEDIA($C$3:$C$54))/CONT.VALORES($C$3:$C$54)"),
        ("Folha top 10",       "=SOMARPRODUTO(GRANDE($C$3:$C$54,{1,2,3,4,5,6,7,8,9,10}))"),
    ]
    for i, (label, form) in enumerate(metricas, start=3):
        ws.cell(row=i, column=6, value=label).font = Font(size=10)
        cell = ws.cell(row=i, column=7, value=form)
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.font = Font(size=10)
        if "%" in label:
            cell.number_format = "0.0%"
        elif label not in ["Desvio Padrao"]:
            cell.number_format = 'R$ #,##0.00'

    apply_table_border(ws, 2, n, 1, 4)
    apply_table_border(ws, 2, 10, 6, 7)

    print("  [OK] Aba 'Metricas' criada")


# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    print("Gerando cheatsheet_excel.xlsx ...")
    wb = Workbook()

    create_dados(wb)
    create_filtros(wb)
    create_agrupamento(wb)
    create_datas(wb)
    create_condicional(wb)
    create_lookup(wb)
    create_metricas(wb)

    wb.save(OUTPUT_FILE)
    size_kb = os.path.getsize(OUTPUT_FILE) / 1024
    print(f"\n[OK] Arquivo salvo: {OUTPUT_FILE}")
    print(f"     Tamanho: {size_kb:.1f} KB")
    print(f"     Abas: {[ws.title for ws in wb.worksheets]}")


if __name__ == "__main__":
    main()
